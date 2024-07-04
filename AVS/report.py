from io import BytesIO

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, PatternFill, Side


def export(data: list):
    excel_file = BytesIO()

    report_data = pd.DataFrame(data)
    report_data.rename(columns={"courseCode": "Course Code", "itRelated": "IT Related"}, inplace=True)
    report_data["IT Related"] = report_data["IT Related"].apply(lambda x: "Yes" if x else "No")

    report_data.to_excel(excel_file, index=False, engine='openpyxl')

    excel_file.seek(0)

    work_book = load_workbook(excel_file)
    work_sheet = work_book.active

    grey_fill = PatternFill(start_color='FFF2F2F2',
                            end_color='FFF2F2F2',
                            fill_type='solid')

    for cell in work_sheet[1]:
        cell.fill = grey_fill

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for row in work_sheet.iter_rows():
        for cell in row:
            cell.border = thin_border

    excel_file.seek(0)
    work_book.save(excel_file)
    excel_file.seek(0)

    return excel_file
